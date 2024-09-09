import requests
from datetime import datetime, date, timedelta
import pandas as pd
import base64
import hashlib
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import styleframe
import os
import sys
from openpyxl.styles import PatternFill, Font, Color
import glob
import pymysql
from sqlalchemy import create_engine, Table, MetaData, Column, String, inspect



print('begin')

host = 'xxxxxx'
port = 3307
user = 'DATeam'
password = 'xxxxxx'
database = 'fdl'

date_variable = datetime.today().date()
keyword = "人事資料表"
file_paths = glob.glob("Z:/18_各部門共享區/01_會計課/03_薪資相關/08_人事資料表(共用)/*{}*.xlsx".format(keyword))
if file_paths:
    file_path = file_paths[0]
    # 讀取 Excel 文件
    depart_df = pd.read_excel(file_path)
depart_super = depart_df[['員工工號', '單位', '部門']]
depart_base = depart_df[['員工工號', '部門', '課別', '組別']]

'''
login
'''
def fn_login(CompanyID, pwd):
    IP_adrs = "xxxxxx"
    url_login = f"http://{IP_adrs}/SCSWeb/api/systemobject/"
    headers = {"Content-Type": "application/json"}

    payload = {
        "Action": "Login",
        "Value": {
            "$type": "AIS.Define.TFindInputArgs, AIS.Define",
            "CompanyID": CompanyID,
            "UserID": "API01",
            "Password": pwd,
            "LanguageId": "zh-Tw"
        }
    }

    response = requests.post(url_login, headers = headers, json=payload)
    content = response.json()
    SessionGuid = content['SessionGuid']

    return SessionGuid

'''
day off
'''
def off_data(SessionGuid, CompanyID, start, end):
    IP_adrs = "xxxxxx"
    url_obj = f"http://{IP_adrs}/SCSServices/api/businessobject/"
    
    payload_obj = {
        "Action": "ExecReport",
        "SessionGuid": SessionGuid,
        "ProgID": "RATT025",
        "Value": {
            "$type": "AIS.Define.TExecReportInputArgs, AIS.Define",
            "UIType": "Report",
            "ReportID": "",
            "ReportTailID": "",
            "FilterItems": [
                {
                    "$type": "AIS.Define.TFilterItem, AIS.Define",
                    "FieldName": "A.SYS_CompanyID",
                    "FilterValue": CompanyID
                },
                {
                    "$type": "AIS.Define.TFilterItem, AIS.Define",
                    "FieldName": "C.AttendDate",
                    "FilterValue": f"{start.strftime('%Y/%m/%d')},{end.strftime('%Y/%m/%d')}",
                    "ComparisonOperator": "Between"
                }
            ],
            "UserFilter": ""
        }
    }
    
    response = requests.post(url_obj, json=payload_obj)
    content_obj = response.json()
    off_all  = pd.json_normalize(content_obj["DataSet"]["ReportBody"])
    return off_all


def calculate_start_date(current_date):
    # Extract the year and month from the current date
    year = current_date.year
    month = current_date.month

    # Define the default start date
    start_date = datetime(year, 1, 1).date()

    # Calculate the start date based on the current month
    if month > 1:  # If current month is not January
        start_date = datetime(year, month - 1, 26).date()
    else:
        start_date = datetime(year - 1, month + 11, 26).date()

    return start_date

# Fetch the current date dynamically using datetime.today()
current_date = datetime.today().date()- timedelta(days=1)

# Calculate the start date using the provided current date
start_date = calculate_start_date(current_date)
if current_date >= datetime(current_date.year, current_date.month, 26).date():
    start_date = datetime(current_date.year, current_date.month, 26).date()

    print("Adjusted start date:", start_date)
    print("Start date:", start_date)

end_date = date.today()
yesterday = end_date - timedelta(days=1)  # 这也是一个 date 类型对象

pwd_TW = "xxxxxx"
CompanyID_TW = "SCS001"
SessionGuid_TW = fn_login(CompanyID_TW, pwd_TW)
late_off_all = off_data(SessionGuid_TW, CompanyID_TW, start_date, yesterday)
late_df = late_off_all[['TMP_EMPLOYEEID', 'TMP_EMPLOYEENAME', 'ATTENDDATE', 'WORKTIME', 'LATEMINUTES', 'TMP_ATTLEAVEHOURS']]
late_df = late_df.rename(columns={
    'TMP_EMPLOYEEID': '出勤人員',
    'TMP_EMPLOYEENAME': '人員姓名',
    'ATTENDDATE': '出勤日期',
    'WORKTIME': '出勤時間(B1)',
    'LATEMINUTES': '遲到分鐘數',
    'TMP_ATTLEAVEHOURS': '銷假時數'
})
#late_df = late_df.sort_values(by=['出勤部門', '人員姓名', '出勤日期'])
late_df = late_df.drop_duplicates(subset=['出勤日期', '人員姓名'], keep='last')
late_df = late_df[late_df['遲到分鐘數']!=0]
late_df['銷假時數'] = late_df['銷假時數'].fillna(0)
late_df['銷假時數'] = late_df['銷假時數']*60
final_df = late_df[(late_df['銷假時數'] < late_df['遲到分鐘數'])]
final_df['出勤日期'] = final_df['出勤日期'].str[:10]
final_tw_sorted = final_df.sort_values(by=['人員姓名', '出勤日期'])
final_tw_sorted.drop(columns=['銷假時數'], inplace=True)
#record_df = pd.read_excel('C:/Users/user/Desktop/project/hr_robot/人事表.xlsx')
record_df = depart_df.copy()
#record_df = record_df[~record_df['職務英文名稱'].isna()]
record_df = record_df[record_df['職務'].isin(['副理', '協理', '特助', '資深經理', '經理', '資深副理', '副理', '代副理'])]
#merge
final_super = final_tw_sorted[final_tw_sorted['人員姓名'].isin(record_df['中文姓名'])]
final_base = final_tw_sorted[~final_tw_sorted['人員姓名'].isin(record_df['中文姓名'])]
final_super['出勤人員'] = final_super['出勤人員'].astype(str)
depart_super['員工工號'] = depart_super['員工工號'].astype(str)
final_base['出勤人員'] = final_base['出勤人員'].astype(str)
depart_base['員工工號'] = depart_base['員工工號'].astype(str)
final_super = pd.merge(final_super, depart_super, left_on = '出勤人員', right_on = '員工工號')
final_base = pd.merge(final_base, depart_base, left_on = '出勤人員', right_on = '員工工號')

final_super = final_super.sort_values(by=['單位', '部門', '出勤人員', '出勤日期'])
final_base = final_base.sort_values(by=['部門', '課別', '組別', '出勤人員', '出勤日期'])
final_super.insert(0, '單位', final_super.pop('單位'))
final_super.insert(1, '部門', final_super.pop('部門'))
final_base.insert(0, '部門', final_base.pop('部門'))
final_base.insert(1, '課別', final_base.pop('課別'))
final_base.insert(2, '組別', final_base.pop('組別'))
final_super.drop(columns=['員工工號'], inplace=True)
final_base.drop(columns=['員工工號'], inplace=True)

final_super = final_super.fillna('')
final_base = final_base.fillna('')
final_super.to_excel(f"主管人員遲到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx", index=False)
final_base.to_excel(f"人員遲到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx", index=False)
#############################################################################################################################
'''
supervisor late data
'''
writer = pd.ExcelWriter(f"主管人員遲到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx")
final_super.to_excel(writer, sheet_name='sheetName', index=False, na_rep='NaN')
excel_writer = styleframe.StyleFrame.ExcelWriter(f"主管人員遲到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx")
sf = styleframe.StyleFrame(final_super)
columns = ['出勤人員', '人員姓名', '出勤日期', '出勤時間(B1)', '遲到分鐘數']
num_rows = len(sf)
if num_rows >0:
    columns_and_rows_to_freeze = 'B2'
else:
    columns_and_rows_to_freeze = None
sf.to_excel(excel_writer=excel_writer, best_fit=columns, row_to_add_filters=0,
            columns_and_rows_to_freeze = columns_and_rows_to_freeze)
excel_writer.close()
with pd.ExcelWriter(f"主管人員遲到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx", engine='openpyxl') as writer:
    final_super.to_excel(writer, sheet_name='sheetName', index=False, na_rep='NaN')
# Apply formatting to the Excel file using openpyxl
filename = f"主管人員遲到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx"
workbook = load_workbook(filename)
sheet = workbook.active
sheet.freeze_panes = 'A2'
sheet.column_dimensions['A'].width = 17
sheet.column_dimensions['B'].width = 15
sheet.column_dimensions['C'].width = 15
sheet.column_dimensions['E'].width = 15
sheet.column_dimensions['F'].width = 15
sheet.column_dimensions['G'].width = 15
name_counts = final_super['人員姓名'].value_counts()
duplicate_names = name_counts[name_counts > 1].index
red_fill = PatternFill(start_color='FFFFDDCC', end_color='FFFFDDCC', fill_type='solid')

for row in sheet.iter_rows(min_row=2, min_col=1, max_col=7):
    if row[3].value in duplicate_names:  # Check the '人員姓名' column
        for cell in row:  # Apply formatting to '出勤日期', '出勤時間(B1)', '人員姓名', '出勤人員' columns
            cell.fill = red_fill
# 新增的資料需改變字體顏色
end_date_str = yesterday.strftime('%Y-%m-%d')  # 將 end_date 轉換為字符串
blue_font = Font(color='FF0000')
for row in sheet.iter_rows(min_row=2, min_col=1, max_col=7):
    cell_date_str = row[4].value
    if cell_date_str == end_date_str:  # 比較字符串形式的日期
        for cell in row:  # 對 '出勤日期', '出勤時間(B1)', '人員姓名', '出勤人員' 欄位應用格式
            cell.font = blue_font            
sheet.auto_filter.ref = sheet.dimensions
workbook.save(filename)
workbook.close()

'''
base late data
'''
writer = pd.ExcelWriter(f"人員遲到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx") 
final_base.to_excel(writer, sheet_name='sheetName', index=False, na_rep='NaN')
excel_writer = styleframe.StyleFrame.ExcelWriter(f"人員遲到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx")
sf = styleframe.StyleFrame(final_base)
columns = ['出勤人員', '人員姓名', '出勤日期', '出勤時間(B1)', '遲到分鐘數']
num_rows = len(sf)
if num_rows >0:
    columns_and_rows_to_freeze = 'B2'
else:
    columns_and_rows_to_freeze = None
sf.to_excel(excel_writer=excel_writer, best_fit=columns, row_to_add_filters=0,
            columns_and_rows_to_freeze = columns_and_rows_to_freeze)
excel_writer.close()

with pd.ExcelWriter(f"人員遲到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx", engine='openpyxl') as writer:
    final_base.to_excel(writer, sheet_name='sheetName', index=False, na_rep='NaN')
# Apply formatting to the Excel file using openpyxl
filename = f"人員遲到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx"
workbook = load_workbook(filename)
sheet = workbook.active
sheet.freeze_panes = 'A2'
sheet.column_dimensions['A'].width = 17
sheet.column_dimensions['B'].width = 15
sheet.column_dimensions['C'].width = 15
sheet.column_dimensions['D'].width = 15
sheet.column_dimensions['E'].width = 15
sheet.column_dimensions['F'].width = 15
sheet.column_dimensions['G'].width = 15
sheet.column_dimensions['H'].width = 15
name_counts = final_base['人員姓名'].value_counts()
duplicate_names = name_counts[name_counts > 1].index
red_fill = PatternFill(start_color='FFFFDDCC', end_color='FFFFDDCC', fill_type='solid')

for row in sheet.iter_rows(min_row=2, min_col=1, max_col=8):
    if row[4].value in duplicate_names:  # Check the '人員姓名' column
        for cell in row:  # Apply formatting to '出勤日期', '出勤時間(B1)', '人員姓名', '出勤人員' columns
            cell.fill = red_fill
            
end_date_str = yesterday.strftime('%Y-%m-%d')  # 將 end_date 轉換為字符串
blue_font = Font(color='FF0000')
for row in sheet.iter_rows(min_row=2, min_col=1, max_col=8):
    cell_date_str = row[5].value
    if cell_date_str == end_date_str:  # 比較字符串形式的日期
        for cell in row:  # 對 '出勤日期', '出勤時間(B1)', '人員姓名', '出勤人員' 欄位應用格式
            cell.font = blue_font              
sheet.auto_filter.ref = sheet.dimensions
workbook.save(filename)
workbook.close()
'''
to 企微
'''
def upload_file_super(file):
    url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key=bot_key&type=file"  # 機器人的Webhook
    headers = {"Content-Type": "multipart/form-data"}
    data = {'file':open(file, 'rb')}
    response = requests.post(url, headers=headers,files=data)
    content = response.json()
    media_id = content['media_id']
    wx_url = 'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=bot_key'
    data = {'msgtype':'file', 'file':{'media_id':media_id}}
    response2 = requests.post(url = wx_url, json = data)
    content2 = response2.json()
    return content2

def text_base(file):
    url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=bot_key"
    headers = {"Content-Type": "application/json"}
    current_date = yesterday.strftime('%Y-%m-%d')
    late_records = final_base[final_base['出勤日期'] == current_date][['部門', '課別', '人員姓名']]   
    content = f'遲到日期: <font color="warning">{current_date}</font>，遲到資訊請相關同事注意\n'
    if not late_records.empty:
        content += ">遲到部門:\n"
        for (department, team), group in late_records.groupby(['部門', '課別']):
            count = len(group)  # 获取当前部门和课别的人数
            persons = ', '.join(group['人員姓名'])  # 获取当前部门和课别下的所有人名
            content += f"<font color='comment'>{department} {team}: <font color='red'>{count}</font> {persons}</font>\n"
   # 构建完整的数据
    data = {
        "msgtype": "markdown",
        "markdown": {
            "content": content
        }
    }
    r = requests.post(url, headers=headers, json=data)
    return r

def upload_file_base(file):
    url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key=bot_key&type=file"  # 機器人的Webhook
    headers = {"Content-Type": "multipart/form-data"}
    data = {'file':open(file, 'rb')}
    response = requests.post(url, headers=headers,files=data)
    content = response.json()
    media_id = content['media_id']
    wx_url = 'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=bot_key'
    data = {'msgtype':'file', 'file':{'media_id':media_id}}
    response2 = requests.post(url = wx_url, json = data)
    content2 = response2.json()
    return content2

#############################################海外
def oversea_login(CompanyID, pwd):
    IP_adrs = "192.168.1.131"
    url_login = f"http://{IP_adrs}/SCSWeb/api/systemobject/"
    headers = {"Content-Type": "application/json"}

    payload = {
        "Action": "Login",
        "Value": {
            "$type": "AIS.Define.TFindInputArgs, AIS.Define",
            "CompanyID": CompanyID,
            "UserID": "API01",
            "Password": pwd,
            "LanguageId": "zh-Tw"
        }
    }

    response = requests.post(url_login, headers = headers, json=payload)
    content = response.json()
    SessionGuid_ov = content['SessionGuid']

    return SessionGuid_ov

'''
day off
'''
def oversea_off_data(SessionGuid, CompanyID, start, end):
    IP_adrs = "xxxxxx"
    url_obj = f"http://{IP_adrs}/SCSServices/api/businessobject/"
    
    payload_obj = {
        "Action": "ExecReport",
        "SessionGuid": SessionGuid,
        "ProgID": "RATT025",
        "Value": {
            "$type": "AIS.Define.TExecReportInputArgs, AIS.Define",
            "UIType": "Report",
            "ReportID": "",
            "ReportTailID": "",
            "FilterItems": [
                {
                    "$type": "AIS.Define.TFilterItem, AIS.Define",
                    "FieldName": "A.SYS_CompanyID",
                    "FilterValue": CompanyID
                },
                {
                    "$type": "AIS.Define.TFilterItem, AIS.Define",
                    "FieldName": "C.AttendDate",
                    "FilterValue": f"{start.strftime('%Y/%m/%d')},{end.strftime('%Y/%m/%d')}",
                    "ComparisonOperator": "Between"
                }
            ],
            "UserFilter": ""
        }
    }
    
    response = requests.post(url_obj, json=payload_obj)
    content_obj = response.json()
    off_all  = pd.json_normalize(content_obj["DataSet"]["ReportBody"])
    return off_all

start_date = calculate_start_date(current_date)
if current_date >= datetime(current_date.year, current_date.month, 26).date():
    start_date = datetime(current_date.year, current_date.month, 26).date()

    print("Adjusted start date:", start_date)
    print("Start date:", start_date)
end_date = date.today()
pwd_OV = "xxxxxx"
CompanyID_OV = "SCS002"
SessionGuid_OV = oversea_login(CompanyID_OV, pwd_OV)
late_off_all = oversea_off_data(SessionGuid_OV, CompanyID_OV, start_date, yesterday)
late_df = late_off_all[['TMP_PROFITNAME', 'TMP_EMPLOYEEID', 'TMP_EMPLOYEENAME', 'ATTENDDATE', 'WORKTIME', 'LATEMINUTES', 'TMP_ATTLEAVEHOURS']]
late_df = late_df.rename(columns={
    'TMP_PROFITNAME': '部門名稱',
    'TMP_EMPLOYEEID': '出勤人員',
    'TMP_EMPLOYEENAME': '人員姓名',
    'ATTENDDATE': '出勤日期',
    'WORKTIME': '出勤時間(B1)',
    'LATEMINUTES': '遲到分鐘數',
    'TMP_ATTLEAVEHOURS': '銷假時數'
})
late_df = late_df.sort_values(by=['部門名稱', '人員姓名', '出勤日期'])
late_df = late_df.drop_duplicates(subset=['出勤日期', '人員姓名'], keep='last')
late_df = late_df[late_df['遲到分鐘數']!=0]
late_df['銷假時數'] = late_df['銷假時數'].fillna(0)
late_df['銷假時數'] = late_df['銷假時數']*60
final_df = late_df[(late_df['銷假時數'] < late_df['遲到分鐘數'])]
final_df['出勤日期'] = final_df['出勤日期'].str[:10]
final_ov_sorted = final_df.sort_values(by=['部門名稱', '人員姓名', '出勤日期'])
final_ov_sorted.drop(columns=['銷假時數'], inplace=True)
final_ov_sorted.to_excel(f"海外人員遲到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx", index=False)
#################################################################################################################################
writer = pd.ExcelWriter(f"海外人員遲到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx") 
final_ov_sorted.to_excel(writer, sheet_name='sheetName', index=False, na_rep='NaN')
excel_writer = styleframe.StyleFrame.ExcelWriter(f"海外人員遲到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx")
sf = styleframe.StyleFrame(final_ov_sorted)
columns = ['部門名稱', '出勤人員', '人員姓名', '出勤日期', '出勤時間(B1)', '遲到分鐘數']
num_rows = len(sf)
if num_rows >0:
    columns_and_rows_to_freeze = 'B2'
else:
    columns_and_rows_to_freeze = None
sf.to_excel(excel_writer=excel_writer, best_fit=columns, row_to_add_filters=0,
            columns_and_rows_to_freeze = columns_and_rows_to_freeze)
excel_writer.close()

with pd.ExcelWriter(f"海外人員遲到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx", engine='openpyxl') as writer:
    final_ov_sorted.to_excel(writer, sheet_name='sheetName', index=False, na_rep='NaN')

# Apply formatting to the Excel file using openpyxl
filename = f"海外人員遲到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx"
workbook = load_workbook(filename)
sheet = workbook.active
sheet.freeze_panes = 'A2'
sheet.column_dimensions['A'].width = 17
sheet.column_dimensions['B'].width = 15
sheet.column_dimensions['D'].width = 15
sheet.column_dimensions['E'].width = 15
sheet.column_dimensions['F'].width = 15
name_counts = final_ov_sorted['人員姓名'].value_counts()
duplicate_names = name_counts[name_counts > 1].index
red_fill = PatternFill(start_color='FFFFDDCC', end_color='FFFFDDCC', fill_type='solid')

for row in sheet.iter_rows(min_row=2, min_col=1, max_col=7):
    if row[2].value in duplicate_names:  # Check the '人員姓名' column
        for cell in row[:-1]:  # Apply formatting to '出勤日期', '出勤時間(B1)', '人員姓名', '出勤人員' columns
            cell.fill = red_fill
end_date_str = yesterday.strftime('%Y-%m-%d')  # 將 end_date 轉換為字符串
blue_font = Font(color='FF0000')
for row in sheet.iter_rows(min_row=2, min_col=1, max_col=7):
    cell_date_str = row[3].value
    if cell_date_str == end_date_str: 
        for cell in row:  # 對 '出勤日期', '出勤時間(B1)', '人員姓名', '出勤人員' 欄位應用格式
            cell.font = blue_font      
sheet.auto_filter.ref = sheet.dimensions
workbook.save(filename)
workbook.close()


'''
to 企微
'''
def upload_ov_file(file):
    url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key=bot_key&type=file"  # 機器人的Webhook
    headers = {"Content-Type": "multipart/form-data"}
    data = {'file':open(file, 'rb')}
    response = requests.post(url, headers=headers,files=data)
    content = response.json()
    media_id = content['media_id']
    wx_url = 'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=bot_key'
    data = {'msgtype':'file', 'file':{'media_id':media_id}}
    response2 = requests.post(url = wx_url, json = data)
    content2 = response2.json()
    return content2

def text_ov(file):
    url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=bot_key"
    headers = {"Content-Type": "application/json"}
    current_date = yesterday.strftime('%Y-%m-%d')
    late_records = file[file['出勤日期'] == current_date][['部門名稱', '人員姓名']]   
    content = f'遲到日期: <font color="warning">{current_date}</font>，遲到資訊請相關同事注意\n'
    if not late_records.empty:
        content += ">遲到部門:\n"
        for department, group in late_records.groupby(['部門名稱']):
            count = len(group)  # 获取当前部门和课别的人数
            persons = ', '.join(group['人員姓名'])
            #department = department[0]      
            content += f"<font color='comment'>{department}: <font color='red'>{count}</font> {persons}</font>\n"
            print(content)
    content
    # 构建完整的数据
    data = {
        "msgtype": "markdown",
        "markdown": {
            "content": content
        }
    }
    r = requests.post(url, headers=headers, json=data)
    return r
    #text_ov(final_ov_sorted)
    #text_base(final_base)
    
############################################################################大陸
def china_login(CompanyID, pwd):
    IP_adrs = "192.168.1.131"
    url_login = f"http://{IP_adrs}/SCSWeb/api/systemobject/"
    headers = {"Content-Type": "application/json"}

    payload = {
        "Action": "Login",
        "Value": {
            "$type": "AIS.Define.TFindInputArgs, AIS.Define",
            "CompanyID": CompanyID,
            "UserID": "API01",
            "Password": pwd,
            "LanguageId": "zh-Tw"
        }
    }

    response = requests.post(url_login, headers = headers, json=payload)
    content = response.json()
    SessionGuid_cn = content['SessionGuid']

    return SessionGuid_cn

'''
day off
'''
def china_off_data(SessionGuid, CompanyID, start, end):
    IP_adrs = "xxxxxx"
    url_obj = f"http://{IP_adrs}/SCSServices/api/businessobject/"
    
    payload_obj = {
        "Action": "ExecReport",
        "SessionGuid": SessionGuid,
        "ProgID": "RATT025",
        "Value": {
            "$type": "AIS.Define.TExecReportInputArgs, AIS.Define",
            "UIType": "Report",
            "ReportID": "",
            "ReportTailID": "",
            "FilterItems": [
                {
                    "$type": "AIS.Define.TFilterItem, AIS.Define",
                    "FieldName": "A.SYS_CompanyID",
                    "FilterValue": CompanyID
                },
                {
                    "$type": "AIS.Define.TFilterItem, AIS.Define",
                    "FieldName": "C.AttendDate",
                    "FilterValue": f"{start.strftime('%Y/%m/%d')},{end.strftime('%Y/%m/%d')}",
                    "ComparisonOperator": "Between"
                }
            ],
            "UserFilter": ""
        }
    }
    
    response = requests.post(url_obj, json=payload_obj)
    content_obj = response.json()
    off_all  = pd.json_normalize(content_obj["DataSet"]["ReportBody"])
    return off_all

start_date = calculate_start_date(current_date)
if current_date >= datetime(current_date.year, current_date.month, 26).date():
    start_date = datetime(current_date.year, current_date.month, 26).date()

    print("Adjusted start date:", start_date)
    print("Start date:", start_date)
end_date = date.today()
pwd_CN = "xxxxxx"
CompanyID_CN = "SCS003"
SessionGuid_CN = china_login(CompanyID_CN, pwd_CN)
late_off_all = china_off_data(SessionGuid_CN, CompanyID_CN, start_date, yesterday)
late_df = late_off_all[['TMP_PROFITNAME', 'TMP_EMPLOYEEID', 'TMP_EMPLOYEENAME', 'ATTENDDATE', 'WORKTIME', 'LATEMINUTES', 'TMP_ATTLEAVEHOURS']]
late_df = late_df.rename(columns={
    'TMP_PROFITNAME': '部门名称',
    'TMP_EMPLOYEEID': '出勤人员',
    'TMP_EMPLOYEENAME': '人员姓名',
    'ATTENDDATE': '出勤日期',
    'WORKTIME': '出勤时间(B1)',
    'LATEMINUTES': '迟到分钟数',
    'TMP_ATTLEAVEHOURS': '销假时数'
})
late_df = late_df.sort_values(by=['部门名称', '人员姓名', '出勤日期'])
late_df = late_df.drop_duplicates(subset=['出勤日期', '人员姓名'], keep='last')
late_df = late_df[late_df['迟到分钟数']!=0]
late_df['销假时数'] = late_df['销假时数'].fillna(0)
late_df['销假时数'] = late_df['销假时数']*60
final_df = late_df[(late_df['销假时数'] < late_df['迟到分钟数'])]
final_df['出勤日期'] = final_df['出勤日期'].str[:10]
final_cn_sorted = final_df.sort_values(by=['部门名称', '人员姓名', '出勤日期'])
final_cn_sorted.drop(columns=['销假时数'], inplace=True)
final_cn_sorted.to_excel(f"大陆人员迟到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx", index=False)
writer = pd.ExcelWriter(f"大陆人员迟到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx") 
final_cn_sorted.to_excel(writer, sheet_name='sheetName', index=False, na_rep='NaN')
excel_writer = styleframe.StyleFrame.ExcelWriter(f"大陆人员迟到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx")
sf = styleframe.StyleFrame(final_cn_sorted)
columns = ['部门名称', '出勤人员', '人员姓名', '出勤日期', '出勤时间(B1)', '迟到分钟数']
num_rows = len(sf)
if num_rows >0:
    columns_and_rows_to_freeze = 'B2'
else:
    columns_and_rows_to_freeze = None
sf.to_excel(excel_writer=excel_writer, best_fit=columns, row_to_add_filters=0,
            columns_and_rows_to_freeze = columns_and_rows_to_freeze)
excel_writer.close()

with pd.ExcelWriter(f"大陆人员迟到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx", engine='openpyxl') as writer:
    final_cn_sorted.to_excel(writer, sheet_name='sheetName', index=False, na_rep='NaN')

# Apply formatting to the Excel file using openpyxl
filename = f"大陆人员迟到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx"
workbook = load_workbook(filename)
sheet = workbook.active
sheet.freeze_panes = 'A2'
sheet.column_dimensions['A'].width = 17
sheet.column_dimensions['B'].width = 15
sheet.column_dimensions['D'].width = 15
sheet.column_dimensions['E'].width = 15
sheet.column_dimensions['F'].width = 15
name_counts = final_cn_sorted['人员姓名'].value_counts()
duplicate_names = name_counts[name_counts > 1].index
red_fill = PatternFill(start_color='FFFFDDCC', end_color='FFFFDDCC', fill_type='solid')

for row in sheet.iter_rows(min_row=2, min_col=1, max_col=7):
    if row[2].value in duplicate_names:  # Check the '人員姓名' column
        for cell in row[:-1]:  # Apply formatting to '出勤日期', '出勤時間(B1)', '人員姓名', '出勤人員' columns
            cell.fill = red_fill
end_date_str = yesterday.strftime('%Y-%m-%d')  # 將 end_date 轉換為字符串
blue_font = Font(color='FF0000')
for row in sheet.iter_rows(min_row=2, min_col=1, max_col=7):
    cell_date_str = row[3].value
    if cell_date_str == end_date_str: 
        for cell in row:  # 對 '出勤日期', '出勤時間(B1)', '人員姓名', '出勤人員' 欄位應用格式
            cell.font = blue_font      
sheet.auto_filter.ref = sheet.dimensions
workbook.save(filename)
workbook.close()

'''
to 企微
'''
def upload_cn_file(file):
    url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key=bot_key&type=file"  # 機器人的Webhook
    headers = {"Content-Type": "multipart/form-data"}
    data = {'file':open(file, 'rb')}
    response = requests.post(url, headers=headers,files=data)
    content = response.json()
    media_id = content['media_id']
    wx_url = 'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=bot_key'
    data = {'msgtype':'file', 'file':{'media_id':media_id}}
    response2 = requests.post(url = wx_url, json = data)
    content2 = response2.json()
    return content2

def text_cn(file):
    url = "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=bot_key"
    headers = {"Content-Type": "application/json"}
    #current_date = date.today().strftime('%Y-%m-%d')
    late_records = file[file['出勤日期'] == end_date_str][['部门名称', '人员姓名']]   
    content = f'迟到日期: <font color="warning">{end_date_str}</font>，迟到资讯请相关同事注意\n'
    if not late_records.empty:
        content += ">迟到部门:\n"
        for department, group in late_records.groupby(['部门名称']):
            count = len(group)  # 获取当前部门和课别的人数
            persons = ', '.join(group['人员姓名'])
            #department = department[0]      
            content += f"<font color='comment'>{department}: <font color='red'>{count}</font> {persons}</font>\n"
            print(content)
    content
    # 构建完整的数据
    data = {
        "msgtype": "markdown",
        "markdown": {
            "content": content
        }
    }
    r = requests.post(url, headers=headers, json=data)
    return r
    #text_ov(final_ov_sorted)
    #text_base(final_base)
######################################################

'''
decide to upload or not
'''
#end_date_str = end_date.strftime('%Y-%m-%d')
end_date_str = yesterday.strftime('%Y-%m-%d') 

if end_date_str in final_super['出勤日期'].values: 
    upload_file_super(f"主管人員遲到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx")
if end_date_str in final_base['出勤日期'].values:
    text_base(final_base)
    upload_file_base(f"人員遲到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx")
if end_date_str in final_ov_sorted['出勤日期'].values: 
    text_ov(final_ov_sorted)
    upload_ov_file(f"海外人員遲到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx")
if end_date_str in final_cn_sorted['出勤日期'].values: 
    text_cn(final_cn_sorted)
    upload_cn_file(f"大陆人员迟到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx")
       

os.remove(f"主管人員遲到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx")
os.remove(f"人員遲到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx")
os.remove(f"海外人員遲到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx")
os.remove(f"大陆人员迟到_{start_date.strftime('%m%d')}-{yesterday.strftime('%m%d')}.xlsx")

sys.exit()
